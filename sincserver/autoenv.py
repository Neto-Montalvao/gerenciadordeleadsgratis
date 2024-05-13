def enviasincserver(df):
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import datetime
    import locale
    import requests
    import json
    import bd
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    from email.mime.text import MIMEText
    from datetime import timedelta
    import re
    from urllib.parse import quote
    enviar = False
    locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')

    data_atual = datetime.datetime.now()

    dt = data_atual.strftime('%d/%m/%Y - %H:%M:%S')

    print(dt)


    envios = [
                {'nome': 'neto', 'tell': '5514997603977', 'msg': 'enviandolead', 'imgc': 'vazia', 'fecha':True},
                {'nome': 'neto', 'tell': '5514997603977', 'msg': 'enviandolead', 'imgc': 'vazia', 'fecha':True}
            ]

    def send_email(subject, body, to):
        from_address = "sac@montalvaoimoveis.com.br"
        password = ""

        msg = MIMEMultipart()
        msg['From'] = from_address
        msg['To'] = to
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_address, password)
        text = msg.as_string()
        server.sendmail(from_address, to, text)
        server.quit()
    

    filename = f"allleads.xlsx"

    def corretor(tell, email):
        try:
            tell = apenas_numeros(tell)
            dftmp = pd.read_excel(filename)
            dftmp['tell'] = dftmp['tell'].apply(apenas_numeros)
            ocorr = dftmp[(dftmp['tell'].str.contains(f'{tell}'))|(dftmp['email'] == f'{email}')]
            string = ocorr.iloc[0]['corretor']
        except:
            string = f'Não encontrado'
        return string
    
    def ficha(tell, email):

        dftmp = pd.read_excel(filename)
        ocorr = dftmp[(dftmp['tell'] == tell )|(dftmp['email'] == email)]
        string = ''
        for counter, (i, r) in enumerate(reversed(list(ocorr.iterrows()))):
            string +=f'''
                ------------
                Ocorrência: {counter}
                    Nome: {r["nome"]}
                    
                    Telefone: {r["tell"]}

                    E-mail: {r["email"]}
                    
                    Corretor: {r['corretor']}

                    Data da requisição: {r['data de cadastro']}

                    Data do repasse: {r['data de repasse']}
                ------------

                '''
        return string



    def apenas_numeros(string):
        return re.sub(r'\D', '', f'{string}')
    
    df['Telefone'] = df['Telefone'].apply(apenas_numeros)


    cor = {
        'Carol':{'tell':'5514982321350', 'msg':'', 'g':'a'},
        'Generozo':{'tell':'5514997941529', 'msg':'', 'g':'o'},
        'Joao':{'tell':'5514981635715', 'msg':'', 'g':'o'},
        'Luiza':{'tell':'5514998079089', 'msg':'', 'g':'a'},
        'Paula':{'tell':'5514996258420', 'msg':'', 'g':'a'},
        'Rosana':{'tell':'5514991937006', 'msg':'', 'g':'a'}
        }




    dfd = df.copy()
    dfd['Data de Cadastro'] = pd.to_datetime(dfd['Data de Cadastro'], format='%d/%m/%Y - %H:%M:%S')

    dfd.sort_values('Data de Cadastro', ascending=False, inplace=True)

    df_duplicates = dfd[dfd.duplicated(subset=['E-mail'], keep=False) | dfd.duplicated(subset=['Telefone'], keep=False)]

    # Cria uma máscara booleana para 'E-mail'
    df_duplicates = df_duplicates.groupby('E-mail').filter(lambda x: (x['Data de Cadastro'].max() - x['Data de Cadastro'].min()).days > 2)

    # Cria uma máscara booleana para 'Telefone'
    df_duplicates = df_duplicates.groupby('Telefone').filter(lambda x: (x['Data de Cadastro'].max() - x['Data de Cadastro'].min()).days > 2)


    df_duplicates.to_excel('contatos_duplicados.xlsx')

    df_duplicates = df_duplicates[df_duplicates['ID'] > int(bd.get('last'))]

    for i, r in df_duplicates.iterrows():
        if int(r['ID']) > int(bd.get('last')):
            enviar = True
            send_email(f"RECONTATO DO {r['ID']}", 'Urgente', "neto@montalvaoimoveis.com.br")
            corretorant = corretor(r["Telefone"], r["E-mail"])

            if corretorant == 'em analise':
                tb = { 
                    'repetido': 'Sim',
                    'data de cadastro': r['Data de Cadastro'].strftime('%d/%m/%Y - %H:%M:%S'),
                    'data de repasse': dt,
                    'corretor': 'em analise',
                    'id': r['ID'],
                    'nome': r['Nome'],
                    'tell': str(r['Telefone']),
                    'email': r['E-mail'],
                    'midia de origem': r['Mídia de origem'],
                    'conteudo': r['Conteúdo'],
                }
                
                allleads = pd.read_excel(filename)

                df_cnc = pd.concat([pd.DataFrame.from_dict([tb]), allleads], ignore_index=True)

                df_cnc.to_excel(filename, index=False)

                book = load_workbook(filename)
                sheet = book.active
                for column_cells in sheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length+5

                book.save(filename)
                book.save('todosclientes.xlsx')

                envios.append({'nome': 'Neto', 'tell': '5514997603977', 'msg': f'''Recontato de:

                            ID: {r["ID"]} 
                               
                            Nome: {r["Nome"]}
                            
                            Telefone: {r["Telefone"]} 

                            E-mail: {r["E-mail"]} 
                            
                            Corretor Encarregado: {corretorant}

                                Ficha: [
                                    {ficha(r["Telefone"], r["E-mail"])}
                                ]

                                '''.replace('                            ', '    '), 'imgc': 'vazia', 'fecha':True})
                envios.append({'nome': 'Montalvao Imoveis', 'tell': '5514996892957', 'msg': f'''Tem um cliente em análise pedindo recontato...\nVerifique a conversa com ele e envie para o Neto.\n Nome: {r["Nome"]}\n Telefone: {re.sub(r'\D', '', r["Telefone"])}''', 'imgc': 'vazia', 'fecha':True})
            
            elif not 'Não encontrado' in corretorant:

                tb = { 
                    'repetido': 'Sim',
                    'data de cadastro': r['Data de Cadastro'].strftime('%d/%m/%Y - %H:%M:%S'),
                    'data de repasse': dt,
                    'corretor': 'em analise',
                    'id': r['ID'],
                    'nome': r['Nome'],
                    'tell': str(r['Telefone']),
                    'email': r['E-mail'],
                    'midia de origem': r['Mídia de origem'],
                    'conteudo': r['Conteúdo'],
                }
                
                allleads = pd.read_excel(filename)

                df_cnc = pd.concat([pd.DataFrame.from_dict([tb]), allleads], ignore_index=True)

                df_cnc.to_excel(filename, index=False)

                book = load_workbook(filename)
                sheet = book.active
                for column_cells in sheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length+5

                book.save(filename)
                book.save('todosclientes.xlsx')

                envios.append({'nome': 'Neto', 'tell': '5514997603977', 'msg': f'''Recontato de:

                            ID: {r["ID"]} 
                               
                            Nome: {r["Nome"]}
                            
                            Telefone: {r["Telefone"]} 

                            E-mail: {r["E-mail"]} 
                            
                            Corretor Encarregado: {corretorant}

                                Ficha: [
                                    {ficha(r["Telefone"], r["E-mail"])}
                                ]

                                '''.replace('                            ', '    '), 'imgc': 'vazia', 'fecha':True})
                msg = f'''Oi {r["Nome"].split(' ')[0]}, tudo bem? Esperamos que sim!\n\nRecebemos seus dados de contato solicitando informações sobre o Villagio Real lll.\n\nObservamos em nosso registro que já houve atendimento anteriormente, através d{cor[corretorant]['g']} corretor{'a' if cor[corretorant]['g']=='a' else ''} {corretorant} ({cor[corretorant]['tell']}).\n\nDeseja prosseguir com o corretor anterior ou falar com um novo?'''
                envios.append({'nome': 'Montalvao Imoveis', 'tell': '5514996892957', 'msg': f'''Tivemos um recontato de {r["Nome"]} - ID: {r['ID']}\n\n
                Lembre-se de reportar o resultado por escrito ao Neto em 48 horas.\n\n Clique no link para enviar a mensagem: https://api.whatsapp.com/send?phone={re.sub(r'\D', '', r["Telefone"])}&text={quote(msg)}
                ''', 'imgc': 'vazia', 'fecha':True})
            
            else:
                tb = { 
                    'repetido': 'Sim',
                    'data de cadastro': r['Data de Cadastro'].strftime('%d/%m/%Y - %H:%M:%S'),
                    'data de repasse': dt,
                    'corretor': 'em analise',
                    'id': r['ID'],
                    'nome': r['Nome'],
                    'tell': str(r['Telefone']),
                    'email': r['E-mail'],
                    'midia de origem': r['Mídia de origem'],
                    'conteudo': r['Conteúdo'],
                }
                allleads = pd.read_excel(filename)

                df_cnc = pd.concat([pd.DataFrame.from_dict([tb]), allleads], ignore_index=True)

                df_cnc.to_excel(filename, index=False)

                book = load_workbook(filename)
                sheet = book.active
                for column_cells in sheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length+5
                book.save(filename)
                book.save('todosclientes.xlsx')

                envios.append({'nome': 'Neto', 'tell': '5514997603977', 'msg': f"Não encontramos o {r['ID']}.\n Nome:{r["Nome"]}\n Telefone:{r["Telefone"]}\n E-mail:{ r["E-mail"]}\n {corretorant}", 'imgc': 'vazia', 'fecha':True})


    df_sem_duplicados = df.drop_duplicates(subset=['E-mail'], keep='last')

    df_sem_duplicados = df_sem_duplicados.drop_duplicates(subset=['Telefone'], keep='last')

    df_sem_duplicados.to_excel('contatos_sem_duplicados.xlsx', index=False)

    df_filtrado = df_sem_duplicados[df_sem_duplicados['ID'] > int(bd.get('last'))]
    
    bd.set('last', f'{df.loc[0, "ID"]}')

    print(df_filtrado)

    if not df_filtrado.empty:
        enviar = True
        dfs = {}

        #d = 0
        ci = eval(bd.get('ci'))
        for i, row in df_filtrado.iterrows():            

            corretor = ci.pop(0)
            ci.append(corretor)
            
            tb = { 
                    'repetido': 'Não',
                    'data de cadastro': row['Data de Cadastro'],
                    'data de repasse': dt,
                    'corretor': corretor,
                    'id': row['ID'],
                    'nome': row['Nome'],
                    'tell': row['Telefone'],
                    'email': row['E-mail'],
                    'midia de origem': row['Mídia de origem'],
                    'conteudo': row['Conteúdo'],
                }
            
            
            if corretor not in dfs:
                dfs[corretor] = pd.DataFrame([tb])
            else:
                dfs[corretor] = dfs[corretor]._append(tb, ignore_index=True)



        bd.set('ci', f'{ci}')
        

        df_concat = pd.concat(dfs.values(), ignore_index=True)

        df_existing = pd.read_excel(filename)

        df_final = pd.concat([df_concat, df_existing], ignore_index=True)

        df_final.to_excel(filename, index=False)

        book = load_workbook(filename)
        sheet = book.active
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length+5

        book.save(filename)
        book.save('todosclientes.xlsx')

        print(dfs)


        ctts = []
        
        
        for corretor, df in dfs.items():
            cor[corretor]['msg'] = f"\n\n*LEADS AUTOMÁTICOS PARA _{corretor}_*\n\n"
            
            for i, row in df.iterrows():
                numero = re.sub(r'\D', '', row['tell'])
                cor[corretor]['msg'] += f"{i+1} - {row['nome']} https://wa.me/{numero} {row['email']};\n\n\n"
                ctts.append({'corretor': corretor, 'idcr': f'{i+1}', 'id': row['id'], 'nome': row['nome'], 'tell': row['tell']},)
            cor[corretor]['msg'] += '\n (obs: você tem menos de 48 horas para entrar em contato. Para melhores resultados na conversa, priorize conversar com o cliente o quanto antes!)'
            envios.append({'nome': corretor, 'tell': '+'+cor[corretor]['tell'], 'msg': cor[corretor]['msg'], 'imgc': 'vazia', 'fecha':True})
            send_email(f"ENVIO DE LEADS PARA {corretor}", cor[corretor]['msg'], "neto@montalvaoimoveis.com.br")
        print(ctts)

    else:
        print('\n\n\n\ndataframe vazio\n\n\n\n')


    if enviar:

        url = 'http://localhost:5000/data'
        headers = {'Content-Type': 'application/json'}

        


        data = json.dumps(envios)

        response = requests.post(url, headers=headers, data=data)

        if response.status_code == 200:
            print(response.json())
        else:
            print(f'Error: {response.status_code}')